from LMSdataBackend import mainDb_Create
from LMSdataBackend import schoolClass_CRUD
from LMSdataBackend import profile_CRUD
from LMSdataBackend import schoolDaysPerMonth_CRUD
from LMSdataBackend import gradeJHS_CRUD
from LMSdataBackend import gradeSHS_CRUD
from LMSdataBackend import semesterSubjects_CRUD
from LMSdataBackend import attendance_CRUD
from LMSdataBackend import observedValues_CRUD

import os
from absPath import resource_path
from openpyxl import load_workbook
from openpyxl.styles import Font


class ReportCardExport:
    def __init__(self):
        self.activeClass = mainDb_Create.viewDbNames(1)[0]
        self.className = self.activeClass[1]
        self.classData = schoolClass_CRUD.viewClassData(self.className)[0]
        self.schoolMonthData = schoolDaysPerMonth_CRUD.viewDaysPerMonthData(self.className, 1)[0]
        self.schoolDaysData = schoolDaysPerMonth_CRUD.viewDaysPerMonthData(self.className, 2)[0]  # No. of school days per month
        self.attendanceData = attendance_CRUD.viewAttendanceData(self.className)
        self.ovData = observedValues_CRUD.viewObValData(self.className)

        if self.classData[1] == 'JUNIOR HIGH SCHOOL':
            self.cardTemplate = load_workbook(resource_path("templates\\card_temp_JHSver2.xlsx"))
            self.jhsGrades = gradeJHS_CRUD.viewGradesJHSData(self.className)
        else:
            self.shsGradesSem1 = gradeSHS_CRUD.viewGradesSHSData(self.className, 1)
            self.shsGradesSem2 = gradeSHS_CRUD.viewGradesSHSData(self.className, 2)
            self.shsSubjectSem1 = semesterSubjects_CRUD.viewSemSubjects(self.className, 1)
            self.shsSubjectSem2 = semesterSubjects_CRUD.viewSemSubjects(self.className, 2)
            self.cardTemplate = load_workbook(resource_path("templates\\card_temp_SHSver2.xlsx"))

        self.frontTemplate = self.cardTemplate['Front']
        self.backTemplate = self.cardTemplate['Back']
        self.cardFolderName = f"G{self.classData[5]}-{self.classData[6]}_Report_Card_{self.classData[2]}"

        self.learner1 = ''
        self.learner2 = ''
        self.learner3 = ''

    # <editor-fold desc="data to excel">
    def classDataToExcel(self):
        for cellColumn1 in range(4, 61, 28):
            self.frontTemplate.cell(row=14, column=cellColumn1, value=self.classData[5])  # entries for grade (3x) (3 cells) - (row = 14)
        for cellColumn2 in range(14, 71, 28):
            self.frontTemplate.cell(row=14, column=cellColumn2, value=self.classData[6])  # - entries for section (3x) (3 cells) - (row = 14)

        if self.classData[1] == 'SENIOR HIGH SCHOOL':
            for cellColumn3 in range(6, 63, 28):
                self.frontTemplate.cell(row=15, column=cellColumn3, value=self.classData[8])  # entries for Strand (3x) (3 cells) - (row = 15)
                self.frontTemplate.cell(row=16, column=cellColumn3, value=self.classData[2])  # entries for School Year (3x) (3 cells) - (row = 16)
        else:
            for cellColumn3 in range(6, 63, 28):
                self.frontTemplate.cell(row=15, column=cellColumn3, value=self.classData[2])  # entries for School Year (3x) (3 cells) - (row = 15)

        for cellColumn4 in range(1, 58, 28):
            self.frontTemplate.cell(row=47, column=cellColumn4, value=self.classData[7])   # - entries for Adviser (3x) (3 cells) - (row = 47)
        for cellColumn5 in range(14, 71, 28):
            self.frontTemplate.cell(row=49, column=cellColumn5, value=self.classData[4])  # - entries for School Head (3x) (3 cells) - (row = 49)

    def monthDataToExcel(self):
        monthDatas = self.schoolMonthData[1:-1]
        schoolDaysDatas = self.schoolDaysData[1:]
        for c in range(11, 76, 28):
            for monthColumn, month in enumerate(monthDatas):
                self.backTemplate.cell(row=29, column=(monthColumn + c), value=month)

            for schoolDaysColumn, schoolDays in enumerate(schoolDaysDatas):
                if schoolDays is None:
                    schoolDays = ""
                self.backTemplate.cell(row=33, column=(schoolDaysColumn + c), value=schoolDays)

    def attendanceToExcel(self, learnerId,  columnAtt):
        learnerAttendance = [item for item in self.attendanceData if item[0] == learnerId][0]
        if len(learnerAttendance) > 0:
            for c_pres, present in enumerate(learnerAttendance[5:18]):  # Number of Days Present
                self.backTemplate.cell(row=34, column=(c_pres + columnAtt), value=present)

            for c_abs, absent in enumerate(learnerAttendance[18:]):  # Number of Days Absent
                self.backTemplate.cell(row=35, column=(c_abs + columnAtt), value=absent)

    def valuesToExcel(self, learnerId, columnValues):
        learnerValues = [item for item in self.ovData if item[0] == learnerId][0]
        rowValues = 5
        for valuesStep in range(9, 41, 4):  # step 9, 13, 17, 21, 25, 29, 33, 37, 41 (4 column step)
            for c_values, valuesValue in enumerate(learnerValues[(valuesStep - 4):valuesStep]):  # valuesValue start [5:9], [9:13], [13:17], [17:21], [21:25], [29:33], [33:37],[37:41]
                self.backTemplate.cell(row=rowValues, column=(columnValues + (c_values * 2)), value=valuesValue)
            if valuesStep <= 21:
                rowValues += 2  # increment merge row cell by 2 (cell row 5,7,9,11,13)
            else:
                rowValues += 3  # increment merge row cell by 3 (cell row 16,19)

    def shsSubjectToExcel(self):
        if len(self.shsSubjectSem1) > 0:
            sem1SubjCount = 0  # SHS subject counter (only 10 allowed)
            for subjectRow, s1 in enumerate(self.shsSubjectSem1):
                for excelColumn in range(1, 58, 28):
                    if sem1SubjCount < 30:  # SHS subjects total of 30 (10 subjects each student)
                        if len(s1[2]) > 54:  # If text length of subject is more than 54, reduce font size
                            self.frontTemplate.cell(row=(subjectRow + 20), column=excelColumn, value=s1[2])\
                                .font = Font(name='Cambria', size=7)
                        else:
                            self.frontTemplate.cell(row=(subjectRow + 20), column=excelColumn, value=s1[2])

                        sem1SubjCount += 1

        if len(self.shsSubjectSem2) > 0:
            sem2SubjCount = 0  # SHS subject counter (only 10 allowed)
            for subjectRow, s2 in enumerate(self.shsSubjectSem2):
                for excelColumn in range(1, 58, 28):
                    if sem2SubjCount < 30:  # SHS subjects total of 30 (10 subjects each student)
                        if len(s2[2]) > 54:  # If text length of subject is more than 54, reduce font size
                            self.frontTemplate.cell(row=(subjectRow + 34), column=excelColumn, value=s2[2]) \
                                .font = Font(name='Cambria', size=7)
                        else:
                            self.frontTemplate.cell(row=(subjectRow + 34), column=excelColumn, value=s2[2])

                        sem2SubjCount += 1

    def checkShsRowGradeCount(self, shsGrades):
        shsGradesOnly = list(shsGrades)
        if len(shsGradesOnly[6:-3]) > 30:  # if contains more than 10 subjects
            newShsSubjectCount = shsGradesOnly[6:36]  # remove final grades (3) (fix 10 subjects)
            return newShsSubjectCount
        else:
            newShsSubjectCount = shsGradesOnly[6:-3]  # remove final averages (3)
            return newShsSubjectCount

    def shsGradeToExcel(self, learnerId, columnValues):
        sem1Grades = [item for item in self.shsGradesSem1 if item[0] == learnerId]
        if len(sem1Grades) > 0:
            learnerGrades = self.checkShsRowGradeCount(sem1Grades[0])  # contains only grades without final average
            rowCount = int(len(learnerGrades)/3)
            for sem1Row in range(rowCount):
                q1Grade = learnerGrades[sem1Row*3]
                q2Grade = learnerGrades[(sem1Row*3)+1]
                q1q2Ave = learnerGrades[(sem1Row*3)+2]
                self.frontTemplate.cell(row=(sem1Row+20), column=columnValues, value=q1Grade)
                self.frontTemplate.cell(row=(sem1Row+20), column=columnValues+2, value=q2Grade)
                self.frontTemplate.cell(row=(sem1Row+20), column=columnValues+4, value=q1q2Ave)
        else:
            for delRow in range(10):
                self.frontTemplate.cell(row=(delRow + 20), column=columnValues, value="")
                self.frontTemplate.cell(row=(delRow + 20), column=columnValues + 2, value="")
                self.frontTemplate.cell(row=(delRow + 20), column=columnValues + 4, value="")

        sem2Grades = [item for item in self.shsGradesSem2 if item[0] == learnerId]
        if len(sem2Grades) > 0:
            learnerGrades = self.checkShsRowGradeCount(sem2Grades[0])  # contains only grades without final average
            rowCount = int(len(learnerGrades) / 3)
            for sem2Row in range(rowCount):
                q3Grade = learnerGrades[sem2Row * 3]
                q4Grade = learnerGrades[(sem2Row * 3) + 1]
                q3q4Ave = learnerGrades[(sem2Row * 3) + 2]
                self.frontTemplate.cell(row=(sem2Row + 34), column=columnValues, value=q3Grade)
                self.frontTemplate.cell(row=(sem2Row + 34), column=columnValues + 2, value=q4Grade)
                self.frontTemplate.cell(row=(sem2Row + 34), column=columnValues + 4, value=q3q4Ave)
        else:
            for delRow in range(10):
                self.frontTemplate.cell(row=(delRow + 34), column=columnValues, value="")
                self.frontTemplate.cell(row=(delRow + 34), column=columnValues + 2, value="")
                self.frontTemplate.cell(row=(delRow + 34), column=columnValues + 4, value="")

    def jhsGradeToExcel(self, jhsGrade, columnValues):
        jhsRow = 20  # starting row
        for rowStep in range(11, 51, 5):  # Grouping grades in 5 (rows) Fil to MAPEH with Final
            for jhsColumn, jhsValue in enumerate(jhsGrade[(rowStep-5):rowStep]):
                self.frontTemplate.cell(row=jhsRow, column=(columnValues + (jhsColumn * 2)), value=jhsValue)
            if rowStep < 31:  # Increment for row cells of grades (pls refer to excel template) JHS
                jhsRow += 2
            elif rowStep == 31:  # (AP to ESP )increment row
                jhsRow += 3
            elif rowStep > 31:  # ( ESP to TLE to MAPEH ) increment row
                jhsRow += 4

        musicToHealthRow = 40
        for rowStep in range(51, 71, 5):  # Grouping grades in 5 Music to Heath
            for jhsColumn, jhsValue in enumerate(jhsGrade[(rowStep-5):(rowStep-1)]):
                self.frontTemplate.cell(row=musicToHealthRow, column=(columnValues + (jhsColumn * 2)), value=jhsValue)
            musicToHealthRow += 1

    def learnerEntry1(self):  # First student entry
        if len(self.learner1) > 0:
            self.frontTemplate.cell(row=12, column=4, value=self.learner1[1])  # Name entry
            self.frontTemplate.cell(row=12, column=24, value=self.learner1[4])  # Age entry
            self.frontTemplate.cell(row=13, column=4, value=self.learner1[3])  # Sex entry
            self.frontTemplate.cell(row=13, column=14, value=self.learner1[2])  # LRN entry

            self.attendanceToExcel(self.learner1[0], 11)
            self.valuesToExcel(self.learner1[0], 20)

            if self.classData[1] == 'SENIOR HIGH SCHOOL':
                self.shsGradeToExcel(self.learner1[0], 19)
            else:
                self.jhsGradeToExcel(self.learner1, 9)

    def learnerEntry2(self):  # Second student entry
        if len(self.learner2) > 0:
            self.frontTemplate.cell(row=12, column=32, value=self.learner2[1])  # Name entry
            self.frontTemplate.cell(row=12, column=52, value=self.learner2[4])  # Age entry
            self.frontTemplate.cell(row=13, column=32, value=self.learner2[3])  # Sex entry
            self.frontTemplate.cell(row=13, column=42, value=self.learner2[2])  # LRN entry

            self.attendanceToExcel(self.learner2[0], 39)
            self.valuesToExcel(self.learner2[0], 48)

            if self.classData[1] == 'SENIOR HIGH SCHOOL':
                self.shsGradeToExcel(self.learner2[0], 47)
            else:
                self.jhsGradeToExcel(self.learner2, 37)

    def learnerEntry3(self):  # Third student entry
        if len(self.learner3) > 0:
            self.frontTemplate.cell(row=12, column=60, value=self.learner3[1])  # Name entry
            self.frontTemplate.cell(row=12, column=80, value=self.learner3[4])  # Age entry
            self.frontTemplate.cell(row=13, column=60, value=self.learner3[3])  # Sex entry
            self.frontTemplate.cell(row=13, column=70, value=self.learner3[2])  # LRN entry

            self.attendanceToExcel(self.learner3[0], 67)
            self.valuesToExcel(self.learner3[0], 76)

            if self.classData[1] == 'SENIOR HIGH SCHOOL':
                self.shsGradeToExcel(self.learner3[0], 75)
            else:
                self.jhsGradeToExcel(self.learner3, 65)

    def learnerDelete2(self):  # Deleting second student entry
        self.frontTemplate.cell(row=12, column=32, value="")
        self.frontTemplate.cell(row=12, column=52, value="")  # deleting name,age,sex,LRN
        self.frontTemplate.cell(row=13, column=32, value="")
        self.frontTemplate.cell(row=13, column=42, value="")

        for attDel2 in range(13):
            self.backTemplate.cell(row=33, column=(attDel2 + 39), value="")
            self.backTemplate.cell(row=34, column=(attDel2 + 39), value="")  # deleting attendance
            self.backTemplate.cell(row=35, column=(attDel2 + 39), value="")

        for rowDel in range(5, 14, 2):   # deleting Observe Values
            for columnDel in range(48, 55, 2):
                self.backTemplate.cell(row=rowDel, column=columnDel, value="")

        for rowDel2 in range(16, 20, 3):  # deleting Observe Values
            for columnDel in range(48, 55, 2):
                self.backTemplate.cell(row=rowDel2, column=columnDel, value="")

        if self.classData[1] == 'SENIOR HIGH SCHOOL':  # deleting SHS grades
            for rowDel in range(10):
                self.frontTemplate.cell(row=(20 + rowDel), column=47, value="")
                self.frontTemplate.cell(row=(34 + rowDel), column=47, value="")
                self.frontTemplate.cell(row=(20 + rowDel), column=49, value="")
                self.frontTemplate.cell(row=(34 + rowDel), column=49, value="")
                self.frontTemplate.cell(row=(20 + rowDel), column=51, value="")
                self.frontTemplate.cell(row=(34 + rowDel), column=51, value="")
        else:
            jhsRowDel = 20
            for rowStep in range(10, 50, 5):
                for jhsColumnDel in range(5):
                    self.frontTemplate.cell(row=jhsRowDel, column=(37 + (jhsColumnDel * 2)), value='')  # deleting JHS grades
                if rowStep < 30:
                    jhsRowDel += 2
                elif rowStep == 30:
                    jhsRowDel += 3
                elif rowStep > 30:
                    jhsRowDel += 4

            musicHealthDel = 40
            for rowStep in range(50, 70, 5):
                for jhsColumnDel in range(4):
                    self.frontTemplate.cell(row=musicHealthDel, column=(37 + (jhsColumnDel * 2)), value='')

                musicHealthDel += 1

    def learnerDelete3(self):  # Deleting third student entry
        self.frontTemplate.cell(row=12, column=60, value="")
        self.frontTemplate.cell(row=12, column=80, value="")  # deleting name,age,sex,LRN
        self.frontTemplate.cell(row=13, column=60, value="")
        self.frontTemplate.cell(row=13, column=70, value="")

        for attDel3 in range(13):
            self.backTemplate.cell(row=33, column=(attDel3 + 67), value="")
            self.backTemplate.cell(row=34, column=(attDel3 + 67), value="")  # deleting attendance
            self.backTemplate.cell(row=35, column=(attDel3 + 67), value="")

        for rowDel in range(5, 14, 2):
            for columnDel in range(76, 83, 2):  # deleting Observe Values
                self.backTemplate.cell(row=rowDel, column=columnDel, value="")

        for rowDel2 in range(16, 20, 3):
            for columnDel in range(76, 83, 2):
                self.backTemplate.cell(row=rowDel2, column=columnDel, value="")

        if self.classData[1] == 'SENIOR HIGH SCHOOL':  # deleting SHS grades
            for rowDel in range(10):
                self.frontTemplate.cell(row=(20 + rowDel), column=75, value="")
                self.frontTemplate.cell(row=(34 + rowDel), column=75, value="")
                self.frontTemplate.cell(row=(20 + rowDel), column=77, value="")
                self.frontTemplate.cell(row=(34 + rowDel), column=77, value="")
                self.frontTemplate.cell(row=(20 + rowDel), column=79, value="")
                self.frontTemplate.cell(row=(34 + rowDel), column=79, value="")
        else:
            jhsRowDel = 20
            for rowStep in range(10, 50, 5):
                for jhsColumnDel in range(5):
                    self.frontTemplate.cell(row=jhsRowDel, column=(65 + (jhsColumnDel * 2)), value='')  # deleting JHS grades
                if rowStep < 30:
                    jhsRowDel += 2
                elif rowStep == 30:
                    jhsRowDel += 3
                elif rowStep > 30:
                    jhsRowDel += 4

            musicHealthDel = 40
            for rowStep in range(50, 70, 5):
                for jhsColumnDel in range(4):
                    self.frontTemplate.cell(row=musicHealthDel, column=(65 + (jhsColumnDel * 2)), value='')

                musicHealthDel += 1
    ...

    # </editor-fold>
    def pathToSave(self, path):
        newPath = os.path.join(path, self.cardFolderName)
        if not os.path.exists(newPath):
            os.makedirs(newPath)

        return newPath

    def groupGradesData(self, data):  # split grades data list into chunks with 3 elements
        for x in range(0, len(data), 3):  # Iteration start from 0 database list index up to length of data, with 3 steps
            chunk = data[x: 3 + x]  # loop is by 3 elements iteration (gradesData[0:3]) // from 0-3, then 4-6, and so on

            if len(chunk) < 3:  # If remaining elements is less than 3, it will be added by '' element
                chunk = chunk + ['' for y in range(3 - len(chunk))]
            yield chunk

    def exportCard(self, data, path):
        groupedDatas = list(self.groupGradesData(data))
        filePath = self.pathToSave(path)
        for groupedData in groupedDatas:  # Iterating/separating list into 3 elements/learners
            count = 0  # counter variable for iterating 3 learners (initial value)
            blankEntry = 0  # counter variable for empty value (no learner's data)
            for gradesData in groupedData:
                count += 1
                match count:
                    case 1:
                        self.learner1 = list(gradesData)  # First learner with entry
                        if len(self.learner1) != 0:
                            # ent1sem2 = sem2List(ent1[0])  # sem2 grade data for SHS
                            self.name1 = self.learner1[1].rsplit(",", 2)
                        else:
                            blankEntry += 1

                    case 2:
                        self.learner2 = list(gradesData)  # Second learner with entry
                        if len(self.learner2) != 0:
                            # ent2sem2 = sem2List(ent2[0])  # sem2 grade data for SHS
                            self.name2 = self.learner2[1].rsplit(",", 2)
                        else:
                            blankEntry += 1

                    case 3:
                        self.learner3 = list(gradesData)  # Third learner with entry
                        if len(self.learner3) != 0:
                            # ent3sem2 = sem2List(ent3[0])  # sem2 grade data for SHS
                            self.name3 = self.learner3[1].rsplit(",", 2)
                        else:
                            blankEntry += 1

            match blankEntry:
                case 0:
                    filename3 = f"{self.name1[0]}_{self.name2[0]}_{self.name3[0]}.xlsx"  # file name for card (3 learners)
                    self.classDataToExcel()
                    self.monthDataToExcel()
                    if self.classData[1] == 'SENIOR HIGH SCHOOL':
                        self.shsSubjectToExcel()

                    self.learnerEntry1()
                    self.learnerEntry2()
                    self.learnerEntry3()
                    self.cardTemplate.save(filePath + "\\" + filename3)

                case 1:
                    filename2 = f"{self.name1[0]}_{self.name2[0]}.xlsx"  # file name for card (2 learners)
                    self.classDataToExcel()
                    self.monthDataToExcel()
                    if self.classData[1] == 'SENIOR HIGH SCHOOL':
                        self.shsSubjectToExcel()

                    self.learnerEntry1()
                    self.learnerEntry2()
                    self.learnerDelete3()  # deleting previous entry (third learner)
                    self.cardTemplate.save(filePath + "\\" + filename2)

                case 2:
                    filename1 = f"{self.name1[0]}.xlsx"   # file name for card (1 learner)
                    self.classDataToExcel()
                    self.monthDataToExcel()
                    if self.classData[1] == 'SENIOR HIGH SCHOOL':
                        self.shsSubjectToExcel()

                    self.learnerEntry1()
                    self.learnerDelete2()  # deleting previous entry (second learner)
                    self.learnerDelete3()  # deleting previous entry (third learner)
                    self.cardTemplate.save(filePath + "\\" + filename1)


# print("Exporting...")
# v = ReportCardExport()
# # sampleData = [(2, 'Alfonso, Felixis T.', '367281902873', 'Male', '19'), (3, 'Alfonso, Reo T.', '5473729930-', 'Male', '21'), (1, 'Canada, Ella marie C.', '2737849039', 'Female', '18'), (4, 'Gonzaga, Jacquie', '112456347893', 'Female', '23'), (5, 'Roberto, Jak P.', '123328918298', 'Male', '23')]
# sampJHS = [(1, 'Targaryen, Viserys V.', '120322524820', 'Male', '23', 1, 81, 84, 86, 84, 83.75, 85, 85, 85, 85, 85.0, 86, 86, 86, 86, 86.0, 87, 87, 88, 87, 87.25, 88, 88, 88, 88, 88.0, 89, 89, 90, 89, 89.25, 90, 90, 90, 90, 90.0, 92, 92, 88, 92, 91.0, 91, 91, 91, 91, 91.0, 92, 92, 90, 92, 91.5, 93, 93, 80, 93, 89.75, 94, 94, 94, 94, 94.0, 87.25, 87.625, 87.625, 87.625, 87.53125), (2, 'Lanister, Jayson T.', '120079845932', 'Male', '22', 2, 88, 85, 82, 87, 85.5, 87, 87, 88, 90, 88.0, 90, 89, 87, 90, 89.0, 86, 88, 87, 82, 85.75, 82, 89, 90, 89, 87.5, 87, 88, 89, 90, 88.5, 83, 90, 86, 87, 86.5, 82, 86, 89, 89, 86.5, 87, 89, 90, 87, 88.25, 80, 86, 87, 90, 85.75, 82, 85, 90, 93, 87.5, 82, 87, 89, 87, 86.25, 85.625, 87.75, 87.25, 88.0, 87.15625), (3, 'Velaryon, Baela', '120092342914', 'Female', '21', 3, 89, 90, 85, 87, 87.75, 76, 70, 78, 72, 74.0, '', 89, '', 88, None, 80, 86, 87, 90, 85.75, '', '', 89, '', None, 90, 90, 89, 84, 88.25, '', '', 88, '', None, None, 88, None, None, None, '', 87, 89, '', None, '', 88, '', '', None, 89, 90, '', 89, None, '', 87, 90, '', None, None, None, None, None, None), (4, 'Mormont, Lyanna S.', '110934920893', 'Female', '19', 4, '', 81, '', '', None, 89, '', 90, '', None, '', 87, '', '', None, 90, 88, '', '', None, '', '', '', '', None, 88, 87, 90, 87, 88.0, '', '', 89, '', None, None, None, 86, None, None, '', '', 87, '', None, 88, '', 87, '', None, '', 90, 81, '', None, '', '', 90, '', None, None, None, None, None, None)]
# samplePath = "C:\\Users\\June\\Documents\\LMSApp_sample_Data"
# v.exportCard(sampJHS, samplePath)
# print("Done. Check directory")
