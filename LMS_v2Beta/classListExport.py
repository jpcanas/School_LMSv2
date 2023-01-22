from openpyxl import load_workbook
from openpyxl.styles import Font
import os

import profileDataBase
import classDataBase
import gradesJHSDataBase
import gradesSHSDataBase1

myClass = classDataBase.viewDataclass()
classData = myClass[0]

gradeSec = (f'{classData[4]} - {classData[5]}')
adviser = classData[6]
if classData[1] == "JUNIOR HIGH SCHOOL":
    nameLearner = gradesJHSDataBase.viewGradeJHS()
else:
    nameLearner = gradesSHSDataBase1.viewGradesem1()

profile = profileDataBase.viewstudentData()
# address = profile[8]
# bday = profile[9]

samplePath = "C:\\LMS_appVer2\\sampleData"
filename = "\\sampleList.xlsx"

bookColumn = 2
startRow = 13

border = Border(left=Side(border_style='thin',color='000000'),     # ---------------------- Apply Border style (color black, thin line) -----------------------
        right=Side(border_style='thin',color='000000'),          
        top=Side(border_style='thin',color='000000'),             
        bottom=Side(border_style='thin',color='000000'))

def genData(sheet, nameColumn, addressCol, bdayColumn, blankCol):

    sheet.cell(column=4, row=9, value=gradeSec)  ##### Grade and Section cell ############
    sheet.cell(column=4, row=10, value=adviser)

    row_length = len(profile)

    for r in range(row_length):         ###### apply the border style #####
        data = [sheet.cell(row=(13 + r), column=nameColumn).coordinate]
        cell = str(data[0])
        sheet[cell].border = border

    if addressCol != 0:
        for r in range(row_length):         ###### apply the border style #####
            data = [sheet.cell(row=(13 + r), column=addressCol).coordinate]
            cell = str(data[0])
            sheet[cell].border = border

    if bdayColumn != 0:
        for r in range(row_length):         ###### apply the border style #####
            data = [sheet.cell(row=(13 + r), column=bdayColumn).coordinate]
            cell = str(data[0])
            sheet[cell].border = border

    if blankCol != 0:
        for r in range(row_length):         ###### apply the border style #####
            data = [sheet.cell(row=(13 + r), column=blankCol).coordinate]
            cell = str(data[0])
            sheet[cell].border = border

def nameAddbday(addressData, bdayData):
    # addressData = True
    # bdayData = True

    nameList = []
    names = []

    for name in nameLearner:        ###### list of learner's name ##########
        names.append(name[1])

        nameList.append(names)
        names = []

    if addressData:                 ###### list of learner's Address ##########
        for bdayAdd in profile:
            nameList[bdayAdd[0]-1].append(bdayAdd[8])

    if bdayData:                    ###### list of learner's Birthday ##########
        for bdayAdd in profile:
            nameList[bdayAdd[0]-1].append(bdayAdd[9])

    return nameList


def columnCount(dataCheck):  ######## PARAMETER => List of 3 item (boolean data) ######## 
    columnNum = 1
    for checked in dataCheck:
        if checked == True:
            columnNum += 1

    match columnNum:
        case 1:         
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList1.xlsx')
            bookColumn = 2
            nameColumn = 2
            addressCol = 5
            bdayColumn = 8

        case 2:
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList2.xlsx')
        case 3:
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList3.xlsx')
        case 4:
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList4.xlsx')

    return book



#             book.save(mypath+ "/" + file_name_class)  #------------------------------------ saving new xlsx file with class info ----

# book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList2.xlsx')
# startColumn1 = 2
# startColumn2 = 5

nameColumn = 2
addressCol = 5
bdayColumn = 8

def listToExcel(nameList, nameColumn, addressCol, bdayColumn):

    for row, name in enumerate(nameList, start = 13): ##### Name Data ####
        active_sheet.cell(column=nameColumn, row=row, value=value)
            # print('2', row, name[0])

    for row, addr in enumerate(nameList, start = 13): ##### Address Data ####
        active_sheet.cell(column=addressCol, row=row, value=value)
            # print('2', row, addr[1])

    for row, bdate in enumerate(nameList, start = 13): ##### Bday Data ####
        active_sheet.cell(column=bdayColumn, row=row, value=value)
            # print('2', row, bdate[2])


def column1(sheet, nameList, nameColumn): ######### function for writing data on excel (Template one column) ######

    genData(sheet, 3, 0, 0, 0)
    for row, name in enumerate(nameList, start = 13): ##### Name Data ####
        sheet.cell(column=nameColumn, row=row, value=value)

def column2(sheet, nameList, column2): ######### function for writing data on excel (Template two column) ######

    genData(sheet, 3, 0, 0, 0)
    column1(sheet, nameList, 2)

    for row, addr in enumerate(nameList, start = 13): ##### Column2 Data ####
        active_sheet.cell(column=column2, row=row, value=value)

def column3(sheet, nameList, column3): ######### function for writing data on excel (Template three column) ######

    genData(sheet, 3, 0, 0, 0)
    column1(sheet, nameList, 2)
    column2(sheet, nameList, 5)

    for row, name in enumerate(nameList, start = 13): ##### Column3 Data ####
        sheet.cell(column=column3, row=row, value=value)

def column4(sheet, nameList): ######### function for writing data on excel (Template four column) ######

    genData(sheet, 3, 0, 0, 0)
    column1(sheet, nameList, 2)
    column2(sheet, nameList, 5)
    column3(sheet, nameList, 7)

    for row, name in enumerate(nameList, start = 13): ##### Column4 Data ####
        sheet.cell(column=9, row=row, value=value)


def classListExp(dataCheck):  ######## PARAMETER => List of 3 item (boolean data) ######## 

    nameList = nameAddbday(dataCheck[0], dataCheck[1]):  ### Create temporary list of name, address or birthday depending if checked ####

    match dataCheck:
        case [True, True, True]: #### Name + [Address, Birthday, Blank] (4 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList4.xlsx')
            startColumn1 = 3

        case [True, True, False]: #### Name + [Address, Birthday] (3 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList3.xlsx')
            startColumn1 = 2
            startColumn2 = 5

        case [True, False, True]: #### Name + [Address, Blank] (3 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList3.xlsx')
            startColumn1 = 2
            startColumn2 = 5
            startColumn3 = 8

        case [False, True, True]: #### Name + [Birthday, Blank] (3 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList3.xlsx')

        case [True, False, False]: #### Name + [Address] (2 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList2.xlsx')
            startColumn1 = 2
            startColumn2 = 5
            startColumn3 = 7
            startColumn4 = 9 

        case [False, True, False]: #### Name + [Birthday] (2 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList3.xlsx')

        case [False, False, True]: #### Name + [Blank] (2 columns) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList3.xlsx')

        case [False, False, False]: #### Name Only (1 column) #######
            book = load_workbook('C:\\LMS_appVer2\\tempFiles\\classList1.xlsx')