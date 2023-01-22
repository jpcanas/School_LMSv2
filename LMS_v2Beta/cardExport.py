import profileDataBase
import classDataBase
import classDataBase
import shsSubject
import gradesJHSDataBase
import gradesSHSDataBase1
import gradesSHSDataBase2
import ovDataBase
import attDataBase

from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# ===================================== MODULE FOR EXPORTING REPORT CARD TO EXCEL TEMPLATE ==============================================================

def classdata_xlsx():  
    
    ################## GENERAL FOR JHS AND SHS CLASS DATA #####################  
    
    if classData[1] == 'SENIOR HIGH SCHOOL':
        for cellcol3 in range(6,63,28):  
            frontmp.cell(row=15, column=cellcol3, value= classData[7])  # ------- entries for Strand (3x) (3 cells) ----------------------- (row = 15)
            frontmp.cell(row=16, column=cellcol3, value= classData[2])  # ------- entries for School Year (3x) (3 cells) ------------------ (row = 16)
    else:
        for cellcol3 in range(6,63,28):  
            frontmp.cell(row=15, column=cellcol3, value= classData[2])  # ------- entries for School Year (3x) (3 cells) ----------------------- (row = 15)

    for cellcol1 in range(4,61,28):  # ------------------------------------- entries for grade (3x) (3 cells) ----------------------- (row = 14)
        frontmp.cell(row=14, column=cellcol1, value= classData[4]) 

    for cellcol2 in range(14,71,28): # ------------------------------------- entries for section (3x) (3 cells) ----------------------- (row = 14)
        frontmp.cell(row=14, column=cellcol2, value= classData[5].upper()) 

    for cellcol4 in range(1,58,28): # -------------------------------------- entries for Adviser (3x) (3 cells) ----------------------- (row = 47)
        frontmp.cell(row=47, column=cellcol4, value= classData[6].upper())

    for cellcol5 in range(14,71,28): # ------------------------------------- entries for School Head (3x) (3 cells) ----------------------- (row = 49)
        frontmp.cell(row=49, column=cellcol5, value= classData[3].upper())

def jhsEnt(grd_ent,grdJHScol):

    jhsrow = 20		########### row for the first subject ##########
    for rowstep in range(10,50,5):  ###### Grouping grades in 5 (rows) Fil to MAPEH with Final ###########
        for jhscol, jhsvalue in enumerate(grd_ent[(rowstep-5):rowstep]):
            if not (jhsvalue.isspace() or jhsvalue==''):
                ijhsvalue = int(jhsvalue)
                frontmp.cell(row=jhsrow, column=(grdJHScol+(jhscol*2)), value= ijhsvalue)
            else:
            	frontmp.cell(row=jhsrow, column=(grdJHScol+(jhscol*2)), value=jhsvalue) #--------- blank entry -------------

        if rowstep < 30:		# -------------- Increment for row cells of grades (pls refer to excel template)---JHS
            jhsrow += 2
        elif rowstep == 30:  #  --- (AP to ESP )increment row ------
        	jhsrow += 3
        elif rowstep > 30:	#  ---( ESP to TLE to MAPEH ) increment row ------
        	jhsrow += 4

    musicHealthRow = 40
    for rowstep in range(50,70,5):  ###### Grouping grades in 4 Music to Heath without Final ###########
        for jhscol, jhsvalue in enumerate(grd_ent[(rowstep-5):(rowstep-1)]):
            if not (jhsvalue.isspace() or jhsvalue==''):
                ijhsvalue = int(jhsvalue)
                frontmp.cell(row=musicHealthRow, column=(grdJHScol+(jhscol*2)), value= ijhsvalue)
            else:
            	frontmp.cell(row=musicHealthRow, column=(grdJHScol+(jhscol*2)), value=jhsvalue)

        musicHealthRow += 1

def shs_mysubject():  
	################### FUNCTION FOR SHS SUBJECTS TO EXCEL (GENERAL FOR SHS) ########
	shsSub = shsSubject.viewSubjSHS()
	actual_sub1 = []      # ----------------------- Counter for subjects with entry (1st sem) -----------------------
	actual_sub2 = []      # ----------------------- Counter for subjects with entry (2nd sem) -----------------------

	if len(shsSub) != 0:  # ---------------- Condition if FIRST sem subject has entry --------------------
		mysub1 = shsSub[0]	
		for r in mysub1[1:]:    #------- Iterating subject database --------------
			if r != '':
				actual_sub1.append(r)

		subcount1 = 0   					#--------- SHS subject counter (only 10 allowed), initial value is zero --------------
		for subrow, s in enumerate(actual_sub1):
			for cellsub in range(1,58,28):  
				if subcount1 < 30:		#--------- SHS subjects total of 30 (10 subjects each student) --------------		
					if len(s) > 54:		# ----------- If text lenght of subject is more than 54, reduce font size -----------
						frontmp.cell(row=(subrow+20), column=cellsub, value= s).font = Font(name='Cambria', size=7)    # --- (font size 7)
					else:   # --- (font size 8)
						frontmp.cell(row=(subrow+20), column=cellsub, value= s)
					
					subcount1 += 1

	if len(shsSub) == 2:	# ---------------- Condition if SECOND sem subject has entry --------------------
		mysub2 = shsSub[1]
		for r2 in mysub2[1:]:    #------- Iterating subject database --------------
			if r2 != '':
				actual_sub2.append(r2)

		subcount2 = 0   					#--------- SHS subject counter (only 10 allowed), initial value is zero --------------
		for subrow2, s2 in enumerate(actual_sub2):
			for cellsub2 in range(1,58,28):  
				if subcount2 < 30:			#--------- SHS subjects total of 30 (10 subjects each student) --------------							
					if len(s2) > 54:		# ----------- If text lenght of subject is more than 54, reduce font size -----------
						frontmp.cell(row=(subrow2+34), column=cellsub2, value= s2).font = Font(name='Cambria', size=7)    # --- (font size 7)
					else:   # --- (font size 8)
						frontmp.cell(row=(subrow2+34), column=cellsub2, value= s2)
					
					subcount2 += 1   


def shsgradesEnt(grd_sem1, grd_sem2, col_1st):   ######### Parameter (list of shs grades from sem1 grades database, first column excel, boolean True if all student selected,False otherwise)
	######### SHS FUNCTION FOR COPYING GRADES (1ST AND 2ND SEM) FROM DATABASE TO EXCEL TEMPLATE (PER LEARNER) ########

    for grdsRow1 in range(10):   				#==================== FIRST SEM GRADES ENTRIES =================
        if not ((grd_sem1[5+(grdsRow1*3)]).isspace() or (grd_sem1[5+(grdsRow1*3)]) == ""):
            q1 = int(grd_sem1[5+(grdsRow1*3)])
        else:
            q1 = ""

        if not ((grd_sem1[6+(grdsRow1*3)]).isspace() or (grd_sem1[6+(grdsRow1*3)]) == ""):
            q2 = int(grd_sem1[6+(grdsRow1*3)])
        else:
            q2 = ""

        if not ((grd_sem1[7+(grdsRow1*3)]).isspace() or (grd_sem1[7+(grdsRow1*3)]) == ""):
            fin1 = int(grd_sem1[7+(grdsRow1*3)])
        else:
            fin1 = ""

        frontmp.cell(row=(grdsRow1+20), column=col_1st, value= q1)		########### Quarter 1 grades (Column 19 excel) ######
        frontmp.cell(row=(grdsRow1+20), column=col_1st+2, value= q2)	########### Quarter 2 grades (Column 21 excel) ######
        frontmp.cell(row=(grdsRow1+20), column=col_1st+4, value= fin1)	########### Final grades (Column 23 excel) ######

    for grdsRow2 in range(10):   				
        if not ((grd_sem2[5+(grdsRow2*3)]).isspace() or (grd_sem2[5+(grdsRow2*3)]) == ""):  
            q3 = int(grd_sem2[5+(grdsRow2*3)])
        else:
            q3 = ""

        if not ((grd_sem2[6+(grdsRow2*3)]).isspace() or (grd_sem2[6+(grdsRow2*3)]) == ""):
            q4 = int(grd_sem2[6+(grdsRow2*3)])
        else:
            q4 = ""

        if not ((grd_sem2[7+(grdsRow2*3)]).isspace() or (grd_sem2[7+(grdsRow2*3)]) == ""):
            fin2 = int(grd_sem2[7+(grdsRow2*3)])
        else:
            fin2 = ""

        frontmp.cell(row=(grdsRow2+34), column=col_1st, value= q3)		########### Quarter 3 grades (Column 19 excel) ######
        frontmp.cell(row=(grdsRow2+34), column=col_1st+2, value= q4)	########### Quarter 4 grades (Column 21 excel) ######
        frontmp.cell(row=(grdsRow2+34), column=col_1st+4, value= fin2)	########### Final grades (Column 23 excel) ######

def month_exp():
	################### Function for writing months data to excel (GENERAL FOR ALL)  ############
	newMonth = []
	month = attDataBase.viewMonth()
	myMonth = month[0]
	month1 = myMonth[0].split(",",2)	############ Isolating Month1 ####
	newMonth.append(month1[0])
	for m in myMonth[1:]:
	    newMonth.append(m)

	for t in range(11,76,28):
		for monthcol, m in enumerate(newMonth):
			backtmp.cell(row=29, column=(monthcol+t), value= m.capitalize())
	
def att_Exp(att_entry,columnAtt):
	########################### Function for exporting attendance data (GENERAL FOR ALL JHS AND SHS) ###########
    lrnlist = att_entry[2]
    at = attDataBase.view_att()

    for att in at:
        for attLRN in att:
            if attLRN == lrnlist:
                entAtt = list(att)
                break
    
    for attcolumn1, att_val1 in enumerate(entAtt[3:15]): ############ Total Number of Days ################
        if not (att_val1.isspace() or att_val1==''):
            iatt_val1 = int(att_val1)
            backtmp.cell(row=33, column=(attcolumn1+columnAtt), value= iatt_val1)
        else:
        	backtmp.cell(row=33, column=(attcolumn1+columnAtt), value= att_val1)

    for attcolumn2, att_val2 in enumerate(entAtt[16:28]): ############ Number of Days Present ################
        if not (att_val2.isspace() or att_val2==''):
            iatt_val2 = int(att_val2)
            backtmp.cell(row=34, column=(attcolumn2+columnAtt), value= iatt_val2)
        else:
        	backtmp.cell(row=34, column=(attcolumn2+columnAtt), value= att_val2)

    for attcolumn3, att_val3 in enumerate(entAtt[29:41]): ############ Number of Days Absent ################
    	if not (att_val3.isspace() or att_val3==''):
    	    iatt_val3 = int(att_val3)
    	    backtmp.cell(row=35, column=(attcolumn3+columnAtt), value= iatt_val3)
    	else:
    		backtmp.cell(row=35, column=(attcolumn3+columnAtt), value= att_val3)

def obVal_Exp(val_entry,valcol):
    ############################### Function for exporting values data (GENERAL FOR ALL JHS AND SHS) ###############
    val_lrn = val_entry[2]
    val = ovDataBase.view_values()

    for ob_val in val:
        for valLRN in ob_val:
            if valLRN == val_lrn:
                entVal = list(ob_val)
                break
    valrow = 5
    for valstep in range(7,39,4): ######## step 7, 11, 15, 19, 23, 27, 31, 35, 39 (4 column step) #######
        for valcolumn, valvalue in enumerate(entVal[(valstep-4):valstep]): ###### valvalue start [3:7], [7:11], [11:15], [15:19], [19:23], [27:31], [31:35],[35:39]
            backtmp.cell(row=valrow, column=(valcol+(valcolumn*2)), value= valvalue)
        if valstep <= 19:
            valrow += 2 		############# increment merge row cell by 2 (cell row 5,7,9,11,13)
        else:
        	valrow += 3 		############# increment merge row cell by 3 (cell row 16,19)

def learner_ent1(entry, entrySem2):   ################ First student entry (GENERAL FOR ALL, filter with condition) #########

    frontmp.cell(row=12, column=4,  value= entry[1])  ############ Name entry #########
    frontmp.cell(row=12, column=24, value= entry[4])  ############ Age entry ##########
    frontmp.cell(row=13, column=4,  value= entry[3])  ############ Sex entry ##########
    frontmp.cell(row=13, column=14, value= entry[2])  ############ LRN entry ##########
    
    att_Exp(entry, 11)
    obVal_Exp(entry, 20)

    if classData[1] == "SENIOR HIGH SCHOOL":
        shsgradesEnt(entry, entrySem2, 19)
    else:
        jhsEnt(entry,9)  		########### column (9,11,13,15) ##############

def learner_ent2(entry, entrySem2):   ################# Second student entry (GENERAL FOR ALL, filter with condition) ################

    frontmp.cell(row=12, column=32,  value= entry[1])  ############ Name entry #########
    frontmp.cell(row=12, column=52,  value= entry[4])  ############ Age entry ##########
    frontmp.cell(row=13, column=32,  value= entry[3])  ############ Sex entry ##########
    frontmp.cell(row=13, column=42,  value= entry[2])  ############ LRN entry ##########
    
    att_Exp(entry, 39)
    obVal_Exp(entry, 48)

    if classData[1] == "SENIOR HIGH SCHOOL":
        shsgradesEnt(entry, entrySem2, 47)
    else:
        jhsEnt(entry,37)  		########### column (37,39,41,43) ##############

def learner_ent3(entry, entrySem2): ################# Third student entry (GENERAL FOR ALL, filter with condition) #################
    
    frontmp.cell(row=12, column=60,  value= entry[1])  ############ Name entry #########
    frontmp.cell(row=12, column=80,  value= entry[4])  ############ Age entry ##########
    frontmp.cell(row=13, column=60,  value= entry[3])  ############ Sex entry ##########
    frontmp.cell(row=13, column=70,  value= entry[2])  ############ LRN entry ##########
    
    att_Exp(entry, 67)
    obVal_Exp(entry, 76)

    if classData[1] == "SENIOR HIGH SCHOOL":
        shsgradesEnt(entry, entrySem2, 75)
    else:
        jhsEnt(entry,65)  		########### column (65,67,69,71) ###############

def learner2_del():   ############## Deleting second student entry (GENERAL FOR ALL, filter with condition) #############
    
    frontmp.cell(row=12, column=32, value=' ')  
    frontmp.cell(row=12, column=52, value=' ')  ######## deleting name,age,sex,LRN #########
    frontmp.cell(row=13, column=32, value=' ')  
    frontmp.cell(row=13, column=42, value=' ') 

    for attdel2 in range(10):											
        backtmp.cell(row=33, column=(attdel2+39), value= '')
        backtmp.cell(row=34, column=(attdel2+39), value= '') ####### deleting attendance ########
        backtmp.cell(row=35, column=(attdel2+39), value= '') 

    for rowdel in range(5,14,2):
        for valcoldel in range(48,55,2):						
            backtmp.cell(row=rowdel, column=valcoldel, value= '')
 
    for rowdel2 in range(16,20,3):							######## deleting Observe Values #########	
        for valcoldel in range(48,55,2):
            backtmp.cell(row=rowdel2, column=valcoldel, value= '')

    if classData[1] == 'SENIOR HIGH SCHOOL':
        for rowdel in range(10):
            frontmp.cell(row=(20+rowdel), column=47, value=' ') 
            frontmp.cell(row=(34+rowdel), column=47, value=' ')  ###### deleting SHS grades ##########
            frontmp.cell(row=(20+rowdel), column=49, value=' ') 
            frontmp.cell(row=(34+rowdel), column=49, value=' ') 
            frontmp.cell(row=(20+rowdel), column=51, value=' ') 
            frontmp.cell(row=(34+rowdel), column=51, value=' ')
    else:
        jhsrowD = 20
        for rowstep in range(10,50,5): 
            for jhscolD in range(4):
                frontmp.cell(row=jhsrowD, column=(37+(jhscolD*2)), value='') 	###### deleting JHS grades #######	 					 			
            if rowstep < 30:		
                jhsrowD += 2
            elif rowstep == 30:  
                jhsrowD += 3
            elif rowstep > 30:			
                jhsrowD += 4

        musicHealthDel = 40
        for rowstep in range(50,70,5):
        	for jhscolD in range(4):
        	    frontmp.cell(row=musicHealthDel, column=(37+(jhscolD*2)), value='')

        	musicHealthDel += 1

def learner3_del():   ############### Deleting third student entry (GENERAL FOR ALL, filter with condition) ################

    frontmp.cell(row=12, column=60, value=' ')  
    frontmp.cell(row=12, column=80, value=' ')  # ----- deleting name,age,sex,LRN ------------------
    frontmp.cell(row=13, column=60, value=' ')  
    frontmp.cell(row=13, column=70, value=' ') 

    for attdel3 in range(10):											
        backtmp.cell(row=33, column=(attdel3+67), value= '')
        backtmp.cell(row=34, column=(attdel3+67), value= '')  # ----- deleting attendance --------------
        backtmp.cell(row=35, column=(attdel3+67), value= '') 

    for rowdel in range(5,14,2):
        for valcoldel in range(76,83,2):						# ----- deleting Observe Values --------------	
            backtmp.cell(row=rowdel, column=valcoldel, value= '')

    for rowdel2 in range(16,20,3):
        for valcoldel in range(76,83,2):
            backtmp.cell(row=rowdel2, column=valcoldel, value= '')
    
    if classData[1] == 'SENIOR HIGH SCHOOL':
        for rowdel in range(10):
            frontmp.cell(row=(20+rowdel), column=75, value=' ') 
            frontmp.cell(row=(34+rowdel), column=75, value=' ')  ###### deleting SHS grades ##########
            frontmp.cell(row=(20+rowdel), column=77, value=' ') 
            frontmp.cell(row=(34+rowdel), column=77, value=' ')
            frontmp.cell(row=(20+rowdel), column=79, value=' ') 
            frontmp.cell(row=(34+rowdel), column=79, value=' ') 
    else:
        jhsrowD = 20
        for rowstep in range(10,50,5): 
            for jhscolD in range(4):
                frontmp.cell(row=jhsrowD, column=(65+(jhscolD*2)), value='') 	###### deleting JHS grades #######	 					 			
            if rowstep < 30:		
                jhsrowD += 2
            elif rowstep == 30:  
                jhsrowD += 3
            elif rowstep > 30:			
                jhsrowD += 4

        musicHealthDel = 40
        for rowstep in range(50,70,5):
        	for jhscolD in range(4):
        	    frontmp.cell(row=musicHealthDel, column=(65+(jhscolD*2)), value='')
        	    
        	musicHealthDel += 1

def list_split(mylist):   # -------------------- Function for spliting database list into chunks with 3 elements (GENERAL FOR ALL) ------------------
	
	for x in range(0,len(mylist),3):   # ----------- Iteration start from 0 database list index upto lenght of database, with 3 steps --------------
		chunk = mylist[x: 3+x]         # ----------- For loop is by 3 elements iteration (  mylist[0:3]  ) // from 0-3, then 4-6, and so on -------  
				
		if len(chunk)<3:				# ---- If remaining elements is less than 3, it will be added by ('') element ---------------
			chunk = chunk + [ ('') for y in range(3 - len(chunk))]
		yield chunk	

def sem2List(idNum):   ########### getting student data from sem2 database (parameter-student id number) (return value - list of student grade sem2) ###### 
    sem2 = gradesSHSDataBase2.viewGradesem2()
    for grade2 in sem2:
    	if grade2[0] == idNum:
    		mygrade2 = list(grade2)
    		break

    return mygrade2

def exportCard(gradeDB, path): ###### Parameter (list of tuples from exportUI),(path of folder to save),(boolean True if all were selected, False otherwise)
      
    global backtmp
    global frontmp
    global classData

    myClass = classDataBase.viewDataclass()
    classData = myClass[0]

    if classData[1] == 'JUNIOR HIGH SCHOOL':
        cardtemp = load_workbook('C:\\LMS_appVer2\\tempFiles\\card_temp_JHSver2.xlsx')
    else:
    	cardtemp = load_workbook('C:\\LMS_appVer2\\tempFiles\\card_temp_SHSver2.xlsx')

    frontmp = cardtemp['Front']
    backtmp = cardtemp['Back']

    grds_folder = f"{classData[4]}-{classData[5]}_REPORT_CARD_({classData[2]})"   #####Folder name for report card #####
    newpath = f"{path}/{grds_folder}"
    
    if not os.path.exists(newpath):
        os.makedirs(newpath)

    s = list(list_split(gradeDB))   ############ New list from database with groupings (3 elements each) ##########

    for chunk3 in s:   # ------------------ Iterating/separating list into 3 elements/learners -----------------
        count = 0  # ------------- counter variable for iterating 3 learners (initial value) --------------
        blnk_ent = 0  # ------------- counter variable for empty value (no learner's data) --------------
        for inn3 in chunk3:  # ------------ Iterating 3 elements of learners individually ----------- ( 3 Iterations)
            count += 1
            match count:
                case 1:
                    ent1 = list(inn3)   # ----------------- First learner with entry --------------------
                    if len(ent1) != 0:
                        ent1sem2 = sem2List(ent1[0])   ###### sem2 grade data for SHS #######
                        s_ent1 = ent1[1].rsplit(",",2)
                    else:
                        blnk_ent += 1

                case 2:
                    ent2 = list(inn3)   # ----------------- Second learner with entry --------------------
                    if len(ent2) != 0:
                        ent2sem2 = sem2List(ent2[0]) ###### sem2 grade data for SHS #######
                        s_ent2 = ent2[1].rsplit(",",2)
                    else:
                        blnk_ent += 1
							
                case 3:
                    ent3 = list(inn3)  # ----------------- Third learner with entry -----------------------
                    if len(ent3) != 0:
                        ent3sem2 = sem2List(ent3[0]) ###### sem2 grade data for SHS #######
                        s_ent3 = ent3[1].rsplit(",",2)
                    else:
                        blnk_ent += 1

        match blnk_ent:
            case 0:
                filename3 =  str(s_ent1[0] + ',' + s_ent2[0] + ',' + s_ent3[0] + '.xlsx') # ----- file name for card (3 learners)----
                month_exp()
                classdata_xlsx()

                if classData[1] == 'SENIOR HIGH SCHOOL':
                    shs_mysubject()

                learner_ent1(ent1, ent1sem2)
                learner_ent2(ent2, ent2sem2)
                learner_ent3(ent3, ent3sem2)
                cardtemp.save(newpath + "/" + filename3)
            
            case 1:
                filename2 =  str(s_ent1[0] + ',' + s_ent2[0]  + '.xlsx') # ----- file name for card (2 learners)----
                month_exp()
                classdata_xlsx()

                if classData[1] == 'SENIOR HIGH SCHOOL':
                    shs_mysubject()

                learner_ent1(ent1, ent1sem2)
                learner_ent2(ent2, ent2sem2)
                learner3_del()      # ------------ deleting previous entry (third learner) --------------------
                cardtemp.save(newpath + "/" + filename2)

            case 2:
                filename1 =  str(s_ent1[0] + '.xlsx') # ----- file name for card (1 learner)----
                month_exp()
                classdata_xlsx()

                if classData[1] == 'SENIOR HIGH SCHOOL':
                    shs_mysubject()

                learner_ent1(ent1, ent1sem2)
                learner2_del()      # ------------ deleting previous entry (secondlearner) --------------------
                learner3_del()      # ------------ deleting previous entry (third learner) --------------------
                cardtemp.save(newpath + "/" + filename1)


# sam = gradesJHSDataBase.viewGradeJHS()
# samp1 = []
# samp1.append(sam[0])
# sem1 = gradesSHSDataBase1.viewGradesem1()
# sem2 = gradesSHSDataBase2.viewGradesem2()
# sampdata1 = [(1, 'BALONDO, JONREY P.', '112688100008', 'MALE', '16 ', '88', '89', '88', '85', '86', '86', '88', '81', '84', '89', '87', '88', '82', '89', '86', '80', '88', '84', '83', '72', '78', '81', '86', '84', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''), (2, 'BARCELONA, JOHN PAUL A. ', '112688100010', 'MALE', '17 ', '89', '89', '89', '80', '87', '84', '80', '86', '83', '83', '89', '86', '87', '88', '88', '88', '85', '86', '82', '86', '84', '87', '82', '84', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')]    #, (3, 'BERAQUIT, FERNANDO L.', '112688100011', 'MALE', '18 ', '87', '87', '87', '85', '86', '86', '86', '88', '87', '82', '85', '84', '90', '86', '88', '88', '83', '86', '89', '87', '88', '87', '80', '84', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')]
# sampdata2 = [(1, 'BALONDO, JONREY P.', '112688100008', 'MALE', '16 ', '88', '80', '84', '87', '95', '91', '85', '90', '88', '90', '92', '91', '82', '92', '87', '83', '82', '82', '84', '83', '84', '86', '84', '85', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''), (2, 'BARCELONA, JOHN PAUL A. ', '112688100010', 'MALE', '17 ', '87', '', '', '', '88', '', '87', '', '', '', '', '', '89', '87', '88', '', '', '', '', '89', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''), (3, 'BERAQUIT, FERNANDO L.', '112688100011', 'MALE', '18 ', '88', '', '', '89', '', '', '86', '', '', '82', '86', '84', '80', '', '', '86', '', '', '87', '', '', '86', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')]
# samplePath = "C:\\LMS_appVer2\\sampleData"
# exportCard(sampdata1, samplePath)
